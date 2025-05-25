import os
import shutil
from pathlib import Path
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func, cast, Float
from sqlalchemy.exc import NoResultFound
from MainProject.app.database.database import SessionLocal
from MainProject.app.models import Subject, Student
from MainProject.app.models.attendance import Attendance
from MainProject.app.models.failure import Failure
from MainProject.app.models.grade import Grade
from MainProject.app.repositories.student_repository import StudentRepository
import openpyxl
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt


class StudentService:

    @staticmethod
    def add_student(session: SessionLocal, name: str, surname: str, pesel: str) -> int:
        return StudentRepository.add_student(session, name, surname, pesel)

    @staticmethod
    def edit_student(session: SessionLocal, student_id: int, name: str = None,
                     surname: str = None, pesel: str = None):
        StudentRepository.edit_student(session, student_id, name, surname, pesel)

    @staticmethod
    def delete_student(session: SessionLocal, student_id: int):
        StudentRepository.delete_student(session, student_id)

    @staticmethod
    def get_student(session: SessionLocal, student_id: int):
        return StudentRepository.get_student(session, student_id)

    @staticmethod
    def get_student_attendances_from_subject(session: SessionLocal, student_id: int, subject_id: int):
        return session.query(Attendance).filter_by(
            student_id=student_id,
            subject_id=subject_id
        ).all()


    @staticmethod
    def get_student_grades_from_subject(session: SessionLocal, student_id: int, subject_id: int):
        return session.query(Grade).filter_by(
            student_id=student_id,
            subject_id=subject_id
        ).all()

    @staticmethod
    def is_student_at_risk(session: SessionLocal, student_id: int, subject_id: int) -> bool:
        failure = (session.query(Failure).filter_by(
            student_id=student_id,
            subject_id=subject_id)
                   .first())
        return failure is not None #0 zdaje 1 zagrożony


    @staticmethod
    def get_average_from_subject(session: SessionLocal, student_id: int, subject_id: int) -> float:
        grades = session.query(Grade).filter_by(student_id=student_id, subject_id=subject_id).all()
        if not grades:
            return 0.0
        total = sum(grade.worth for grade in grades)
        average = total / len(grades)
        return round(average, 2)

    @staticmethod
    def get_total_average(session: SessionLocal, student_id: int) -> float:
        grades = session.query(Grade).filter_by(student_id=student_id).all()
        if not grades:
            return 0.0  # brak ocen
        total = sum(grade.worth for grade in grades)
        average = total / len(grades)
        return round(average, 2)

    @staticmethod
    def generate_grades_report(session: SessionLocal, student_id: int):
        student = StudentService.get_student(session, student_id)
        filename = "Grades.xlsx"
        reports_folder = Path("../Reports")
        student_folder = reports_folder / f"Student_{student_id}_{student.name}_{student.surname}"
        reports_folder.mkdir(parents=True, exist_ok=True)
        student_folder.mkdir(parents=True, exist_ok=True)
        full_path = reports_folder / student_folder / filename

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Grade report student {student_id}"

        light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

        labels = ["Subject", "Average", "Grades"]
        for i, label in enumerate(labels, start=1):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True, size=15, color="132A58")
        cell = ws.cell(row=1, column=1)
        cell.fill = light_blue_fill

        #pobierz jakoś przedmioty
        subjects = (session.query(Subject)
                    .join(Attendance, Attendance.subject_id == Subject.id)
                    .filter(Attendance.student_id == student_id)
                    .distinct()
                    .all())

        if not subjects:
            raise NoResultFound("Uczeń nie ma obecności na żadnym przedmiocie")

        for col, subject in enumerate(subjects, start=2):
            grades = StudentService.get_student_grades_from_subject(session, student_id, subject.id)
            average = StudentService.get_average_from_subject(session, student_id, subject.id)
            if StudentService.is_student_at_risk(session, student_id, subject.id):
                ws.cell(row=1, column=col, value=subject.name).font = Font(bold=True, size=11, color="FF0000")
                ws.cell(row=2, column=col, value=average).font = Font(size=11, color="FF0000")
            else:
                ws.cell(row=1, column=col, value=subject.name).font = Font(bold=True, size=11)
                ws.cell(row=2, column=col, value=average).font = Font(size=11, color="0089FF")
            cell = ws.cell(row=1, column=col)
            cell.fill = light_blue_fill
            for index, grade in enumerate(grades):
                ws.cell(row=3+index, column=col, value=grade.worth).font = Font(size=11)

        for col in ws.columns:
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = 17

        for row in ws.iter_rows():
            row_index = row[0].row
            ws.row_dimensions[row_index].height = 17



        subject_cells = ws[1][1:]
        subjects = [cell.value for cell in subject_cells]

        value_cells = ws[2][1:]
        values = [cell.value for cell in value_cells]

        plt.figure(figsize=(6, 4))
        bars = plt.bar(subjects, values, color='skyblue')
        plt.title("Average per subject")
        plt.ylabel("Average")
        plt.ylim(1, 6)

        for bar in bars:
            yval = bar.get_height()
            plt.text(bar.get_x() + bar.get_width() / 2, yval + 0.1, f'{yval:.2f}', ha='center', va='bottom')

        img_path = student_folder / "avg_per_subject.png"
        plt.savefig(img_path)
        plt.close()

        img = Image(img_path)
        ws.add_image(img, "B9")

        wb.save(str(full_path))
        os.remove(img_path)


    @staticmethod
    def generate_attendance_report(session: SessionLocal, student_id: int):

        student = StudentService.get_student(session, student_id)
        filename = "Attendances.xlsx"
        reports_folder = Path("../Reports")
        student_folder = reports_folder / f"Student_{student_id}_{student.name}_{student.surname}"
        reports_folder.mkdir(parents=True, exist_ok=True)
        student_folder.mkdir(parents=True, exist_ok=True)
        full_path = reports_folder / student_folder / filename

        light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Attendance report student {student_id}"
        labels = Attendance.possible_status_values

        for i, label in enumerate(labels, start=1):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True, size=15, color="132A58")
        cell = ws.cell(row=1, column=1)
        cell.fill = light_blue_fill
        attendances = (session.query(Attendance).filter(Attendance.student_id == student_id).all())
        subjects = (session.query(Subject).join(Attendance, Attendance.subject_id == Subject.id).filter(Attendance.student_id == student_id).distinct().all())

        for col, subject in enumerate(subjects, start = 2):
            ws.cell(row=1, column=col, value = subject.name).font = Font(bold = True, size = 11)
            for i, label in enumerate(labels[1:]):
                attendance_count = sum(1 for attendance in attendances if attendance.status.lower() == label.lower() and attendance.subject_id == subject.id)
                ws.cell(row=2 + i, column=col, value=attendance_count).font = Font(bold=True, size=11)

        for col in ws.columns:
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = 17

        for row in ws.iter_rows():
            row_index = row[0].row
            ws.row_dimensions[row_index].height = 17

        wb.save(str(full_path))


    @staticmethod
    def generate_statistics_for_student(session: SessionLocal, student_id: int):
        student = StudentService.get_student(session, student_id)
        statistics_folder = Path("../Statistics")
        try:
            shutil.rmtree(statistics_folder)
        except FileNotFoundError:
            pass
        statistics_folder.mkdir(parents=True, exist_ok=True)
        full_path = statistics_folder

        grades = session.query(Subject.name, func.avg(Grade.worth)).select_from(Grade).join(Subject, Subject.id == Grade.subject_id).join(Student, Student.id == Grade.student_id).filter(Grade.student_id == student_id).group_by(Subject.name).all()
        subject_count = session.query(func.count(func.distinct(Attendance.subject_id))).filter(Attendance.student_id == student_id).scalar()
        attendences = session.query(Attendance.status, func.count(Attendance.status) / subject_count).filter(Attendance.student_id == student_id).group_by(Attendance.status).all()
        failure_percentage = session.query(func.count(Failure.student_id)).filter(Failure.student_id == student_id).scalar() / subject_count * 100

        if grades:
            subjects, avg_grades = zip(*grades)
            plt.figure(figsize=(8, 5))
            plt.bar(subjects, avg_grades, color='skyblue')
            plt.title(f'{student.name} {student.surname} Average Grades')
            plt.xlabel('Subjects')
            plt.ylabel('Average Grade')
            plt.xticks(rotation=45)
            plt.tight_layout()
            plt.savefig(full_path / "grades.png")
            plt.close()
        else:
            plt.figure(figsize=(8, 5))
            plt.title(f'{student.name} {student.surname} Average Grades')
            plt.text(0.5, 0.5, 'No grade data available', ha='center', va='center', fontsize=12)
            plt.axis('off')
            plt.tight_layout()
            plt.savefig(full_path / "grades.png")
            plt.close()

        if attendences:
            statuses, counts = zip(*attendences)
            plt.figure(figsize=(8, 5))
            plt.bar(statuses, counts, color='lightgreen')
            plt.title(f'{student.name} {student.surname} Attendance Distribution')
            plt.xlabel('Attendance Status')
            plt.ylabel('Average Count per Subject')
            plt.tight_layout()
            plt.savefig(full_path / "attendances.png")
            plt.close()
        else:
            plt.figure(figsize=(8, 5))
            plt.title(f'{student.name} {student.surname} Attendance Distribution')
            plt.text(0.5, 0.5, 'No attendance data available', ha='center', va='center', fontsize=12)
            plt.axis('off')
            plt.tight_layout()
            plt.savefig(full_path / "attendances.png")
            plt.close()

        passed_percentage = 100 - failure_percentage
        labels = ['Failed Subjects', 'Passed Subjects']
        sizes = [failure_percentage, passed_percentage]
        colors = ['salmon', 'lightgreen']
        plt.figure(figsize=(6, 6))
        plt.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', startangle=90)
        plt.title(f'{student.name} {student.surname} Failure Rate')
        plt.axis('equal')
        plt.tight_layout()
        plt.savefig(full_path / "passed.png")
        plt.close()

    @staticmethod
    def generate_statistics_for_everyone(session: SessionLocal):
        statistics_folder = Path("../Statistics")
        try:
            shutil.rmtree(statistics_folder)
        except FileNotFoundError:
            pass
        statistics_folder.mkdir(parents=True, exist_ok=True)
        full_path = statistics_folder

        grades = session.query(
            Subject.name,
            func.avg(cast(Grade.worth, Float))
        ).select_from(Grade).join(Subject, Subject.id == Grade.subject_id).group_by(Subject.name).all()

        student_count = session.query(func.count(Student.id)).scalar()
        if student_count != 0:
            attendances = session.query(Attendance.status, func.count(Attendance.status)).group_by(
                Attendance.status).all()
        else:
            attendances = []

        if student_count != 0:
            failure_percentage = session.query(
                func.count(func.distinct(Failure.student_id))).scalar() / student_count * 100
        else:
            failure_percentage = 0


        if grades:
            subjects, avg_grades = zip(*grades)
            plt.figure(figsize=(8, 5))
            plt.bar(subjects, avg_grades, color='skyblue')
            plt.title('Average Grades per Subject')
            plt.xlabel('Subjects')
            plt.ylabel('Average Grade')
            plt.ylim(1, 6)
            plt.xticks(rotation=45)
            plt.tight_layout()
            plt.savefig(full_path / "grades.png")
            plt.close()

        if attendances:
            statuses, counts = zip(*attendances)
            plt.figure(figsize=(8, 5))
            plt.bar(statuses, counts, color='#504ef2')
            plt.title('Attendance Status Distribution')
            plt.xlabel('Status')
            plt.ylabel('Average Count per Student')
            plt.tight_layout()
            plt.savefig(full_path / "attendances.png")
            plt.close()

        passed_percentage = 100 - failure_percentage
        labels = ['Failed', 'Passed']
        sizes = [failure_percentage, passed_percentage]
        colors = ['salmon', 'lightgreen']
        plt.figure(figsize=(6, 6))
        plt.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', startangle=90)
        plt.title('Students Performance: Fail vs Pass')
        plt.axis('equal')
        plt.tight_layout()
        plt.savefig(full_path / "passed.png")
        plt.close()

    @staticmethod
    def get_student_by_pesel(session, pesel: str):
        return session.query(Student).filter_by(pesel=pesel).first()



